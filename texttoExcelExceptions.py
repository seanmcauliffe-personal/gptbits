import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill

# Read text file
with open("input.txt", "r") as file:
    lines = file.readlines()

# Extract data and create dataframe
data = []
current_exception = None
for line in lines:
    if "Violation report for exception type:" in line:
        current_exception = line.split("type: ")[1].strip()
    elif len(line.split(";")) == 3:
        col_name, col_value, num = line.strip().split(";")
        num = int(num)
        data.append((current_exception, col_name, f"{col_value} ({num})"))

df = pd.DataFrame(data, columns=["Exception Type", "Column Name", "Column Value"])

# Create a color map for exception types
exception_types = df["Exception Type"].unique()
color_map = {exception: PatternFill(start_color=color, end_color=color, fill_type="solid") for exception, color in zip(exception_types, ["FFFF00", "FFA500", "FF0000", "00FF00", "00FFFF", "0000FF", "8A2BE2", "7FFF00"])}

# Write dataframe to Excel
wb = Workbook()
ws = wb.active
for r in dataframe_to_rows(df, index=False, header=True):
    ws.append(r)

# Apply color to cells based on exception type
for row in range(2, len(data) + 2):
    exception_type = ws.cell(row=row, column=1).value
    ws.cell(row=row, column=1).fill = color_map[exception_type]
    ws.cell(row=row, column=2).fill = color_map[exception_type]
    ws.cell(row=row, column=3).fill = color_map[exception_type]

# Save Excel file
wb.save("output.xlsx")
